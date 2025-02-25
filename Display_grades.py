import flet as ft
import openpyxl

EXCEL_FILE = "ClassePrimaire.xlsx"

def fetch_grades(matricule):
    """Fetch grades for a given matricule from the Excel file."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        # Find the column for the given matricule in row 1
        matricule_col = None
        for col in ws.iter_cols(min_row=1, max_row=1):  # Search only in row 1
            cell_value = col[0].value
            # Convert both the cell value and matricule to strings for comparison
            if str(cell_value) == str(matricule):
                matricule_col = col[0].column
                break

        if not matricule_col:
            return None  # Matricule not found

        # Fetch the name and grades
        name = ws.cell(row=2, column=matricule_col).value  # Assuming name is in row 2
        grades = {}
        subjects = ["Maths", "Sciences", "Histoire", "Geo", "ECM", "Francais", "Anglais"]
        for row_num, subject in enumerate(subjects, start=3):  # Grades start from row 3
            grade = ws.cell(row=row_num, column=matricule_col).value
            grades[subject] = grade

        # Calculate the average
        valid_grades = [grade for grade in grades.values() if isinstance(grade, (int, float))]
        if valid_grades:
            grades["Moyenne"] = sum(valid_grades) / len(valid_grades)
        else:
            grades["Moyenne"] = "N/A"

        return {"name": name, "grades": grades}
    except Exception as e:
        print(f"Error fetching grades: {e}")
        return None

def main(page: ft.Page):
    page.title = "View Student Grades"
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER

    # Input field for Matricule
    matricule_field = ft.TextField(label="Matricule", width=300)

    # Result text for displaying grades
    result_text = ft.Text(size=16)

    # Table to display grades
    grades_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("Subject")),
            ft.DataColumn(ft.Text("Grade")),
        ],
        rows=[]
    )

    def view_grades(e):
        matricule = matricule_field.value.strip()
        if not matricule:
            result_text.value = "Please enter a Matricule!"
            result_text.color = ft.colors.RED
            page.update()
            return

        # Fetch grades from Excel
        data = fetch_grades(matricule)
        if not data:
            result_text.value = f"No data found for Matricule: {matricule}"
            result_text.color = ft.colors.RED
            grades_table.rows = []  # Clear the table
        else:
            result_text.value = f"Grades for Matricule: {matricule}, Name: {data['name']}"
            result_text.color = ft.colors.GREEN

            # Populate the table with grades
            rows = []
            for subject, grade in data["grades"].items():
                rows.append(
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(subject)),
                            ft.DataCell(ft.Text(str(grade))),
                        ]
                    )
                )
            grades_table.rows = rows

        page.update()

    # View Grades button
    view_button = ft.ElevatedButton(text="View Grades", on_click=view_grades)

    # Add all widgets to the page
    page.add(
        ft.Column(
            [
                matricule_field,
                view_button,
                result_text,
                grades_table,
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        )
    )

ft.app(target=main)