import flet as ft
import openpyxl
import os

EXCEL_FILE = "ClassePrimaire.xlsx"

def ensure_excel_file():
    """Ensure the Excel file exists with a proper structure."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = " Grades  "
        # Add headers for subjects in column A
        subjects = ["Matricule", "Name", "Maths", "Sciences", "Histoire", "Geo", "ECM", "Francais", "Anglais"]
        for row_num, subject in enumerate(subjects, start=1):
            ws.cell(row=row_num, column=1, value=subject)
        wb.save(EXCEL_FILE)

def is_unique_matricule(matricule):
    """Check if the given Matricule (ID) is unique in the file."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    # Check row 1 (header row) for existing matricules
    existing_ids = [str(cell.value) for cell in ws[1] if cell.column > 2 and cell.value is not None]
    return matricule not in existing_ids

def save_grades(matricule, name, grades):
    """Save the grades under the matricule column."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        # Find the next empty column starting from column C (column 3)
        col_index = 3  # Start from column C
        while ws.cell(row=1, column=col_index).value is not None:
            col_index += 1

        # Add Matricule and Name in the first two rows of the column
        ws.cell(row=1, column=col_index, value=matricule)
        ws.cell(row=2, column=col_index, value=name)

        # Add grades for each subject (convert to integers)
        subjects = ["Maths", "Sciences", "Histoire", "Geo", "ECM", "Francais", "Anglais"]
        for row_num, subject in enumerate(subjects, start=3):
            grade = grades.get(subject)
            if grade.isdigit():  # Ensure the grade is a valid number
                ws.cell(row=row_num, column=col_index, value=int(grade))  # Save as integer
            else:
                ws.cell(row=row_num, column=col_index, value=0)  # Default to 0 if invalid

        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"Error saving grades: {e}")
        return False

def main(page: ft.Page):
    # Set the theme to light mode (white theme)
    page.theme_mode = ft.ThemeMode.LIGHT

    page.title = "Student Grades Entry"
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER

    ensure_excel_file()  # Ensure the Excel file is set up

    # Input fields for Matricule and Name
    matricule_field = ft.TextField(label="Matricule", width=300)
    name_field = ft.TextField(label="Name", width=300)

    # Subjects and grades table
    subjects = [
        "Maths", "Sciences", "Histoire", "Geo", "ECM", "Francais", "Anglais"
    ]
    grade_fields = {}  # To track grade inputs

    # Create table rows for each subject
    table_rows = []
    for subject in subjects:
        grade_field = ft.TextField(
            width=300,
            height=450,
            input_filter=ft.InputFilter(r'^\d*$'),  # Only allow numeric input
            hint_text="0-20"
        )
        grade_fields[subject] = grade_field
        table_rows.append(
            ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(subject)),
                    ft.DataCell(grade_field),
                ]
            )
        )

    grades_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("Subject")),
            ft.DataColumn(ft.Text("Grade")),
        ],
        rows=table_rows
    )

    # Result text for messages
    result_text = ft.Text(size=16)

    def validate_input(value, min_val, max_val, field_name):
        if not value:
            return f"{field_name} is required!"
        try:
            num = int(value)
            if num < min_val or num > max_val:
                return f"{field_name} must be between {min_val} and {max_val}!"
            return None
        except ValueError:
            return f"{field_name} must be a valid number!"

    def submit_data(e):
        matricule = matricule_field.value.strip()
        name = name_field.value.strip()
        errors = []
        grades = {}

        # Validate matricule and name
        if not matricule:
            errors.append("Matricule is required!")
        if not name:
            errors.append("Name is required!")
        elif not is_unique_matricule(matricule):
            errors.append("Matricule already exists! Please enter a unique one.")

        # Validate grades
        for subject in subjects:
            grade_field = grade_fields[subject]
            grade = grade_field.value.strip()
            error = validate_input(grade, 0, 20, subject)
            if error:
                errors.append(error)
                grade_field.border_color = ft.colors.RED
            else:
                grade_field.border_color = None
                grades[subject] = grade

        if errors:
            result_text.value = "\n".join(errors)
            result_text.color = ft.colors.RED
        else:
            # Save grades to Excel
            if save_grades(matricule, name, grades):
                result_text.value = "Grades saved successfully!"
                result_text.color = ft.colors.GREEN
                # Clear inputs
                matricule_field.value = ""
                name_field.value = ""
                for subject in subjects:
                    grade_fields[subject].value = ""
            else:
                result_text.value = "Failed to save grades. Please try again."
                result_text.color = ft.colors.RED

        page.update()

    # Submit button
    submit_button = ft.ElevatedButton(text="Submit Grades", on_click=submit_data)

    # Add all widgets to the page
    page.add(
        ft.Column(
            [
                ft.Row([matricule_field, name_field], alignment=ft.MainAxisAlignment.CENTER),
                grades_table,
                submit_button,
                result_text,
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        )
    )

ft.app(target=main)